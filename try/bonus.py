from openpyxl import load_workbook
import openpyxl

wb=load_workbook('try.xlsx') #读取excel文件
sheet1=wb.get_sheet_by_name('Sheet1') #学生信息工作表
sheet2=wb.get_sheet_by_name('Sheet2') #班级奖金统计
nrows=sheet1.max_row
for Name in range(2,nrows+1):
    Stu_Class=int(sheet1.cell(row=Name,column=3).value) #读取该学生高一上班级信息
    bonus1=sheet2.cell(row=Stu_Class+1,column=2).value+sheet1.cell(row=Name,column=2).value*0.2 #高一上该生奖金
    sheet2.cell(row=Stu_Class+1,column=2,value=bonus1) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=4).value) #读取该学生高一下班级信息
    bonus2=sheet2.cell(row=Stu_Class+1,column=2).value+sheet1.cell(row=Name,column=2).value*0.2 #高一下该生奖金
    sheet2.cell(row=Stu_Class+1,column=2,value=bonus2) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=5).value) #读取该学生高二上班级信息
    bonus3=sheet2.cell(row=Stu_Class+9,column=2).value+sheet1.cell(row=Name,column=2).value*0.3 #高二上该生奖金
    sheet2.cell(row=Stu_Class+9,column=2,value=bonus3) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=6).value) #读取该学生高二下班级信息
    bonus4=sheet2.cell(row=Stu_Class+9,column=2).value+sheet1.cell(row=Name,column=2).value*0.3 #高二下该生奖金
    sheet2.cell(row=Stu_Class+9,column=2,value=bonus4) #累计入班级奖金

wb.save('try1.xlsx')
