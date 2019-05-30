# -*- coding: utf-8 -*-
"""
Created on Wed Mar 20 19:22:53 2019

@author: Administrator
"""
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
#import openpyxl

wb=load_workbook('try.xlsx') #读取excel文件
sheet1=wb['学生信息'] #学生信息工作表
sheet2=wb['班级奖金'] #班级奖金统计
teachers=wb['教师信息']
nrows=sheet1.max_row
for Name in range(2,nrows+1):
    
    Stu_Class=int(sheet1.cell(row=Name,column=3).value) #读取该学生高一上班级信息
    bonus1=sheet2.cell(row=Stu_Class*2,column=3).value+sheet1.cell(row=Name,column=2).value*0.2 #高一上该生奖金
    sheet2.cell(row=Stu_Class*2,column=3,value=bonus1) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=4).value) #读取该学生高一下班级信息
    bonus2=sheet2.cell(row=Stu_Class*2+1,column=3).value+sheet1.cell(row=Name,column=2).value*0.2 #高一下该生奖金
    sheet2.cell(row=Stu_Class*2+1,column=3,value=bonus2) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=5).value) #读取该学生高二上班级信息
    bonus3=sheet2.cell(row=(Stu_Class+8)*2,column=3).value+sheet1.cell(row=Name,column=2).value*0.3 #高二上该生奖金
    sheet2.cell(row=(Stu_Class+8)*2,column=3,value=bonus3) #累计入班级奖金
    
    Stu_Class=int(sheet1.cell(row=Name,column=6).value) #读取该学生高二下班级信息
    bonus4=sheet2.cell(row=(Stu_Class+8)*2+1,column=3).value+sheet1.cell(row=Name,column=2).value*0.3 #高二下该生奖金
    sheet2.cell(row=(Stu_Class+8)*2+1,column=3,value=bonus4) #累计入班级奖金
 
 
sheet3=wb.create_sheet() #此板块新建老师个人奖金工作表
sheet3.title="教师奖金"
sheet3['A1']="姓名"
sheet3['B1']="奖金"
sheet4=wb['DATA']#  教师奖金系数


Teacher=[]
cell_range=teachers['C2':'K33'] #选取所有教师单元格范围


for Row in cell_range:        #遍历教师单元格，统计教师名单并写入"教师奖金"工作表
    for cell in Row:
        if cell.value not in Teacher and cell.value!=None:
            Teacher.append(cell.value)
for i in range(len(Teacher)):
    sheet3.cell(row=i+2,column=1,value=Teacher[i])

for i in range(len(Teacher)):#遍历教师单元格，统计教师奖金并写入"教师奖金"工作表
    money=0
    for Row in cell_range:
        for cell in Row:
            if cell.value==Teacher[i]:
                money+=sheet2.cell(row=cell.row,column=3).value*sheet4.cell(row=cell.row,column=column_index_from_string(cell.column)).value
    sheet3.cell(row=i+2,column=2,value=money)
 
wb.save('try2.xlsx')

